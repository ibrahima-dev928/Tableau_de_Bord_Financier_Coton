# app/routers/dashboard_report.py
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from uuid import UUID
from app.database import get_db
from app import models
from app.utils.security import get_current_user
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch

router = APIRouter()

# ===============================
# GÉNÉRATION EXCEL (avec db passé en paramètre)
# ===============================
def generate_excel(db, campagne, prev_agri, prev_egre, prev_ventes, achats, transformations, ventes, zones_data, total_volume, total_montant, nb_producteurs):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.merge_cells('A1:F1')
    ws1['A1'] = f"RAPPORT DU DASHBOARD - {campagne.libelle.upper()}"
    ws1['A1'].font = Font(size=16, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')
    ws1.append([])
    ws1.append(["Période :", campagne.date_debut.strftime("%d/%m/%Y"), "au", campagne.date_fin.strftime("%d/%m/%Y")])
    ws1.append(["Statut :", "Active" if campagne.est_active else "Inactive"])
    ws1.append([])
    ws1.append(["📊 INDICATEURS CLÉS"])
    ws1.append(["Volume total collecté (kg)", float(total_volume)])
    ws1.append(["Montant total (FCFA)", float(total_montant)])
    ws1.append(["Nombre de producteurs distincts", nb_producteurs])
    ws1.append(["Coût moyen (FCFA/kg)", float(total_montant / total_volume) if total_volume else 0])
    if prev_agri:
        ws1.append([])
        ws1.append(["🌾 PRÉVISIONS AGRICULTURE"])
        ws1.append(["Volume prévu (tonnes)", float(prev_agri.volume_prevu_tonnes)])
        ws1.append(["Prix plancher (FCFA/kg)", float(prev_agri.prix_plancher)])
        ws1.append(["Seuil d'alerte (FCFA/kg)", float(prev_agri.seuil_alerte)])
        ws1.append(["Délai de paiement (jours)", prev_agri.delai_paiement_jours])
    if prev_egre:
        ws1.append([])
        ws1.append(["⚙️ PRÉVISIONS ÉGRENAGE"])
        ws1.append(["Coton graine prévu (tonnes)", float(prev_egre.coton_graine_prevu_tonnes)])
        ws1.append(["Rendement attendu (%)", float(prev_egre.rendement_attendu_pourcent)])
        ws1.append(["Coût transformation estimé (FCFA)", float(prev_egre.cout_transformation_estime)])
    if prev_ventes:
        ws1.append([])
        ws1.append(["📦 PRÉVISIONS VENTES"])
        ws1.append(["Produit", "Volume (t)", "Prix (FCFA/kg)", "Logistique (FCFA)"])
        for pv in prev_ventes:
            ws1.append([pv.produit, float(pv.volume_prevu_tonnes), float(pv.prix_vente_prevu), float(pv.cout_logistique_estime)])
    # Mise en forme résumé
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and 'PRÉVISIONS' in cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
            if cell.value and isinstance(cell.value, str) and 'INDICATEURS' in cell.value:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")

    # Onglet Achats
    ws2 = wb.create_sheet("Achats")
    ws2.append(["Date", "Producteur", "Zone", "Volume (kg)", "Prix (FCFA/kg)", "Montant (FCFA)", "Statut", "Payé"])
    for a in achats:
        prod = db.query(models.Producteur).filter(models.Producteur.id == a.producteur_id).first()
        zone = db.query(models.Zone).filter(models.Zone.id == a.zone_id).first()
        ws2.append([
            a.date_achat.strftime("%d/%m/%Y"),
            f"{prod.nom} {prod.prenom}" if prod else "",
            zone.nom if zone else "",
            float(a.quantite_kg),
            float(a.prix_kg),
            float(a.montant_total),
            a.statut,
            "Oui" if a.paye else "Non"
        ])

    # Onglet Transformations
    ws3 = wb.create_sheet("Transformations")
    ws3.append(["Date", "Usine", "Coton entrant (kg)", "Fibre (kg)", "Graines (kg)", "Coût (FCFA)"])
    for t in transformations:
        usine = db.query(models.Usine).filter(models.Usine.id == t.usine_id).first()
        ws3.append([
            t.date.strftime("%d/%m/%Y"),
            usine.nom if usine else "",
            float(t.qte_coton_graine_kg),
            float(t.qte_fibre_kg),
            float(t.qte_graine_kg),
            float(t.cout_transformation)
        ])

    # Onglet Ventes
    ws4 = wb.create_sheet("Ventes")
    ws4.append(["Date", "Type", "Volume (kg)", "Prix unitaire", "Devise", "Montant total", "Logistique"])
    for v in ventes:
        ws4.append([
            v.date.strftime("%d/%m/%Y"),
            v.type_vente,
            float(v.quantite_kg),
            float(v.prix_unitaire),
            v.devise,
            float(v.montant_total),
            float(v.couts_logistiques)
        ])

    # Onglet Zones
    ws5 = wb.create_sheet("Zones")
    ws5.append(["Zone", "Volume (t)", "Coût moyen (FCFA/kg)"])
    for z in zones_data:
        ws5.append([z.nom, float(z.volume or 0), float(z.cout_moyen or 0)])

    # Style des en-têtes
    for ws in [ws2, ws3, ws4, ws5]:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ===============================
# GÉNÉRATION PDF (avec db passé en paramètre)
# ===============================
def generate_pdf(db, campagne, prev_agri, prev_egre, prev_ventes, achats, transformations, ventes, zones_data, total_volume, total_montant, nb_producteurs):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    heading_style = ParagraphStyle(name='Heading', parent=styles['Heading2'], fontSize=14, spaceAfter=12)
    normal_style = styles['Normal']
    story = []

    story.append(Paragraph(f"<b>RAPPORT DU DASHBOARD</b>", title_style))
    story.append(Paragraph(f"Campagne : {campagne.libelle}", normal_style))
    story.append(Paragraph(f"Période : {campagne.date_debut.strftime('%d/%m/%Y')} au {campagne.date_fin.strftime('%d/%m/%Y')}", normal_style))
    story.append(Spacer(1, 0.3*inch))

    # Indicateurs clés
    story.append(Paragraph("<b>📊 INDICATEURS CLÉS</b>", heading_style))
    data = [
        ["Volume total collecté", f"{float(total_volume):,.2f} kg"],
        ["Montant total", f"{float(total_montant):,.2f} FCFA"],
        ["Producteurs distincts", str(nb_producteurs)],
        ["Coût moyen", f"{float(total_montant / total_volume) if total_volume else 0:,.2f} FCFA/kg"]
    ]
    t = Table(data, colWidths=[2.5*inch, 2.5*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
    ]))
    story.append(t); story.append(Spacer(1, 0.3*inch))

    if prev_agri:
        story.append(Paragraph("<b>🌾 PRÉVISIONS AGRICULTURE</b>", heading_style))
        data = [
            ["Volume prévu", f"{float(prev_agri.volume_prevu_tonnes):,.2f} t"],
            ["Prix plancher", f"{float(prev_agri.prix_plancher):,.2f} FCFA/kg"],
            ["Seuil d'alerte", f"{float(prev_agri.seuil_alerte):,.2f} FCFA/kg"],
            ["Délai de paiement", f"{prev_agri.delai_paiement_jours} jours"]
        ]
        t = Table(data, colWidths=[2.5*inch, 2.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t); story.append(Spacer(1, 0.3*inch))

    if prev_egre:
        story.append(Paragraph("<b>⚙️ PRÉVISIONS ÉGRENAGE</b>", heading_style))
        data = [
            ["Coton graine prévu", f"{float(prev_egre.coton_graine_prevu_tonnes):,.2f} t"],
            ["Rendement attendu", f"{float(prev_egre.rendement_attendu_pourcent):,.2f}%"],
            ["Coût transformation estimé", f"{float(prev_egre.cout_transformation_estime):,.2f} FCFA"]
        ]
        t = Table(data, colWidths=[2.5*inch, 2.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t); story.append(Spacer(1, 0.3*inch))

    if prev_ventes:
        story.append(Paragraph("<b>📦 PRÉVISIONS VENTES</b>", heading_style))
        data = [["Produit", "Volume (t)", "Prix (FCFA/kg)", "Logistique (FCFA)"]]
        for pv in prev_ventes:
            data.append([pv.produit, f"{float(pv.volume_prevu_tonnes):,.2f}", f"{float(pv.prix_vente_prevu):,.2f}", f"{float(pv.cout_logistique_estime):,.2f}"])
        t = Table(data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t); story.append(Spacer(1, 0.3*inch))

    # Achats (20 derniers)
    if achats:
        story.append(Paragraph("<b>📋 DERNIERS ACHATS</b>", heading_style))
        data = [["Date", "Producteur", "Zone", "Volume (kg)", "Prix", "Montant (FCFA)"]]
        for a in achats[:20]:
            prod = db.query(models.Producteur).filter(models.Producteur.id == a.producteur_id).first()
            zone = db.query(models.Zone).filter(models.Zone.id == a.zone_id).first()
            data.append([
                a.date_achat.strftime("%d/%m/%Y"),
                f"{prod.nom} {prod.prenom}" if prod else "",
                zone.nom if zone else "",
                f"{float(a.quantite_kg):,.2f}",
                f"{float(a.prix_kg):,.2f}",
                f"{float(a.montant_total):,.2f}"
            ])
        t = Table(data, colWidths=[0.8*inch, 1.5*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.2*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t); story.append(Spacer(1, 0.3*inch))

    # Zones
    if zones_data:
        story.append(Paragraph("<b>🗺️ COMPARAISON PAR ZONE</b>", heading_style))
        data = [["Zone", "Volume (t)", "Coût moyen (FCFA/kg)"]]
        for z in zones_data:
            data.append([z.nom, f"{float(z.volume or 0):,.2f}", f"{float(z.cout_moyen or 0):,.2f}"])
        t = Table(data, colWidths=[2.0*inch, 2.0*inch, 2.0*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
        ]))
        story.append(t)

    doc.build(story)
    buffer.seek(0)
    return buffer

# ===============================
# ENDPOINT EXPORT
# ===============================
@router.get("/export")
def export_dashboard_report(
    campagne_id: UUID = Query(...),
    format: str = Query("excel", regex="^(excel|pdf)$"),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    campagne = db.query(models.Campagne).filter(models.Campagne.id == campagne_id).first()
    if not campagne:
        raise HTTPException(404, "Campagne non trouvée")

    prev_agri = db.query(models.PrevisionAgriculture).filter_by(campagne_id=campagne_id).first()
    prev_egre = db.query(models.PrevisionEgrenage).filter_by(campagne_id=campagne_id).first()
    prev_ventes = db.query(models.PrevisionVente).filter_by(campagne_id=campagne_id).all()
    achats = db.query(models.Achat).filter(models.Achat.campagne_id == campagne_id).all()
    transformations = db.query(models.Transformation).filter(models.Transformation.campagne_id == campagne_id).all()
    ventes = db.query(models.Vente).filter(models.Vente.campagne_id == campagne_id).all()
    total_volume = db.query(func.sum(models.Achat.quantite_kg)).filter(models.Achat.campagne_id == campagne_id).scalar() or 0
    total_montant = db.query(func.sum(models.Achat.montant_total)).filter(models.Achat.campagne_id == campagne_id).scalar() or 0
    nb_producteurs = db.query(models.Achat.producteur_id).filter(models.Achat.campagne_id == campagne_id).distinct().count()
    zones_data = db.query(
        models.Zone.nom,
        func.sum(models.Achat.quantite_kg).label('volume'),
        func.avg(models.Achat.prix_kg).label('cout_moyen')
    ).join(models.Achat, models.Zone.id == models.Achat.zone_id)\
     .filter(models.Achat.campagne_id == campagne_id)\
     .group_by(models.Zone.nom).all()

    if format.lower() == "excel":
        output = generate_excel(db, campagne, prev_agri, prev_egre, prev_ventes, achats, transformations, ventes, zones_data, total_volume, total_montant, nb_producteurs)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    else:
        output = generate_pdf(db, campagne, prev_agri, prev_egre, prev_ventes, achats, transformations, ventes, zones_data, total_volume, total_montant, nb_producteurs)
        media_type = "application/pdf"
        ext = "pdf"

    filename = f"dashboard_{campagne.libelle.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.{ext}"
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )