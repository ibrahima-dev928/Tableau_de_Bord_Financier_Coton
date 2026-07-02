from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from uuid import UUID
from app.database import get_db
from app import models, schemas
from app.utils.security import get_current_user, require_role
from decimal import Decimal
import io
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter()

@router.post("/", response_model=schemas.AchatResponse, status_code=status.HTTP_201_CREATED)
def create_achat(
    achat: schemas.AchatCreate,
    db: Session = Depends(get_db),
    current_user = Depends(require_role("Responsable_terrain"))
):
    if current_user.zone_id and current_user.zone_id != achat.zone_id:
        raise HTTPException(403, "Vous ne pouvez saisir que dans votre zone")
    
    campagne_active = db.query(models.Campagne).filter(models.Campagne.est_active == True).first()
    campagne_id = campagne_active.id if campagne_active else None
    
    montant_total = achat.quantite_kg * achat.prix_kg
    db_achat = models.Achat(
        **achat.dict(),
        montant_total=montant_total,
        saisi_par_id=current_user.id,
        statut="en_attente",
        campagne_id=campagne_id
    )
    db.add(db_achat)
    db.commit()
    db.refresh(db_achat)
    return db_achat

@router.get("/", response_model=List[schemas.AchatResponse])
def get_achats(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(models.Achat)
    if current_user.role == "Responsable_terrain":
        query = query.filter(models.Achat.zone_id == current_user.zone_id)
    achats = query.offset(skip).limit(limit).all()
    
    result = []
    for a in achats:
        producteur = db.query(models.Producteur).filter(models.Producteur.id == a.producteur_id).first()
        zone = db.query(models.Zone).filter(models.Zone.id == a.zone_id).first()
        saisi_par = db.query(models.User).filter(models.User.id == a.saisi_par_id).first()
        result.append({
            "id": a.id,
            "date_achat": a.date_achat,
            "producteur_id": a.producteur_id,
            "zone_id": a.zone_id,
            "quantite_kg": a.quantite_kg,
            "prix_kg": a.prix_kg,
            "montant_total": a.montant_total,
            "saisi_par_id": a.saisi_par_id,
            "statut": a.statut,
            "producteur_nom": f"{producteur.nom} {producteur.prenom}" if producteur else None,
            "zone_nom": zone.nom if zone else None,
            "saisi_par_nom": saisi_par.nom if saisi_par else None,
        })
    return result

@router.get("/export")
def export_achats_excel(
    zone_id: Optional[str] = Query(None),
    statut: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    query = db.query(models.Achat)
    if zone_id:
        query = query.filter(models.Achat.zone_id == zone_id)
    if statut:
        query = query.filter(models.Achat.statut == statut)

    achats = query.all()
    if not achats:
        # Renvoyer un fichier vide mais valide
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Achats"
        ws.append(["Aucune donnée"])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=achats_vide.xlsx"}
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achats"

    # En-têtes
    headers = ["Date", "Producteur", "Zone", "Volume (kg)", "Prix (FCFA/kg)", "Montant (FCFA)", "Saisi par", "Statut"]
    ws.append(headers)
    
    # Style des en-têtes
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E3D", end_color="1F4E3D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Données
    for a in achats:
        producteur = db.query(models.Producteur).filter(models.Producteur.id == a.producteur_id).first()
        zone = db.query(models.Zone).filter(models.Zone.id == a.zone_id).first()
        saisi_par = db.query(models.User).filter(models.User.id == a.saisi_par_id).first()
        ws.append([
            a.date_achat.strftime("%Y-%m-%d"),
            f"{producteur.nom} {producteur.prenom}" if producteur else "",
            zone.nom if zone else "",
            float(a.quantite_kg),
            float(a.prix_kg),
            float(a.montant_total),
            saisi_par.nom if saisi_par else "",
            a.statut
        ])

    # Ajuster la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 35)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=achats_export.xlsx"}
    )


@router.get("/{achat_id}", response_model=schemas.AchatResponse)
def get_achat(
    achat_id: UUID,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    achat = db.query(models.Achat).filter(models.Achat.id == achat_id).first()
    if not achat:
        raise HTTPException(404, "Achat non trouvé")
    if current_user.role == "Responsable_terrain" and achat.zone_id != current_user.zone_id:
        raise HTTPException(403, "Accès non autorisé")
    return achat

@router.put("/{achat_id}/valider", response_model=schemas.AchatResponse)
def valider_achat(
    achat_id: UUID,
    db: Session = Depends(get_db),
    current_user = Depends(require_role("Comptabilite"))
):
    achat = db.query(models.Achat).filter(models.Achat.id == achat_id).first()
    if not achat:
        raise HTTPException(404, "Achat non trouvé")
    achat.statut = "valide"
    achat.paye = True
    db.commit()
    db.refresh(achat)
    return achat

@router.put("/{achat_id}/rejeter", response_model=schemas.AchatResponse)
def rejeter_achat(
    achat_id: UUID,
    db: Session = Depends(get_db),
    current_user = Depends(require_role("Comptabilite"))
):
    achat = db.query(models.Achat).filter(models.Achat.id == achat_id).first()
    if not achat:
        raise HTTPException(404, "Achat non trouvé")
    achat.statut = "rejete"
    db.commit()
    db.refresh(achat)
    return achat